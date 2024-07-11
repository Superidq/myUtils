"""
对excel表中的特定列的特定数据求和，本实例是对VideosInfo2Excel.py中文本化表示的时间求和
"""
import openpyxl

# 求和文本化时间所在列
sum_column = 'D'


def getNumFromText(text):
    hours, minutes, seconds = text.split(':')
    if hours.isdigit():
        return int(hours), int(minutes), int(seconds)
    return -1


if __name__ == '__main__':
    excel_path = "E:/Downloads/list1.xlsx"
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    sum_hours, sum_minutes, sum_seconds = 0, 0, 0

    # 从第二行开始读取，一般来说都是第二行开始有有效数据，第一行是说明
    for row in range(2, sheet.max_row + 1):
        hours, minutes, seconds = getNumFromText(sheet.cell(row=row, column=ord(sum_column)-ord('A')+1).value)
        sum_hours += hours
        sum_minutes += minutes
        sum_seconds += seconds

    minutes, seconds = divmod(sum_seconds, 60)
    sum_minutes += minutes
    hours, minutes = divmod(sum_minutes, 60)
    sum_hours += hours
    print('求和得：', sum_hours, ':',  minutes, ':', seconds, sep='')

    print("done.")
