"""
将文件夹内批量视频文件信息导入到Excel中，支持在本程序运行生成的Excel文档末尾继续追加新文件信息

From Google Translate:
Import the batch video file information in the folder into Excel,
and support adding new file information at the end of
the Excel document generated by running this program
"""

import os
import shutil
import time

import cv2
import openpyxl.cell
import xlsxwriter
from openpyxl.styles import Alignment
from openpyxl.drawing import image


# 插入图片的列
image_column = 'E'


def TimeStampToTime(timestamp):
    timeStruct = time.localtime(timestamp)
    return time.strftime('%Y-%m-%d %H:%M:%S', timeStruct)


# def get_FileCreateTime(filePath):
#     # '''获取文件的创建时间'''
#     # filePath = unicode(filePath,'utf8')
#     t = os.path.getctime(filePath)
#     return TimeStampToTime(t)


def get_FileModifyTime(filePath):
    """ 获取文件的修改时间"""
    # filePath = unicode(filePath, 'utf8')
    t = os.path.getmtime(filePath)
    return TimeStampToTime(t)


def get_FileSize(filePath):
    """ 获取文件的大小,结果保留两位小数，单位为MB"""
    # filePath = unicode(filePath,'utf8')
    fsize = os.path.getsize(filePath)
    fsize = fsize / float(1024 * 1024)
    return round(fsize, 2)


def get_FileDuration(filepath):
    """ 获取视频时长"""
    cap = cv2.VideoCapture(filepath)
    if cap.isOpened():
        rate = cap.get(5)
        frame_num = cap.get(7)
        duration = frame_num / rate
        return int(duration)
    return -1


def get_SpecifiedFrame(filepath, frame_index, output_path):
    """
    截取选定视频设定的的帧数生成缩略图，以.png格式保存在指定路径
    :param filepath: 视频文件所在位置
    :param frame_index: 设定的截取帧数
    :param output_path: 生成的缩略图保存位置
    :return:
    """
    # 输出文件夹不存在，则创建输出文件夹
    if not os.path.exists(output_path):
        os.mkdir(output_path)

    # 读取视频文件
    cap = cv2.VideoCapture(filepath)
    # 从指定帧开始读取文件
    cap.set(cv2.CAP_PROP_POS_FRAMES, float(frame_index))
    # 用来记录帧
    res, image = cap.read()

    # 等比例缩放(600, 336)->(116, 65), 600/116≈5.17, 116/600≈0.193
    # 实操后发现需调整*1.25，即0.193*1.25=0.241
    # 此处“多此一举”用缩放比例来计算的原因是防止宽高与此不一致的图片缩放后变形
    scale_percent = 24.1
    width = int(image.shape[1] * scale_percent / 100)
    height = int(image.shape[0] * scale_percent / 100)
    image = cv2.resize(image, (width, height), interpolation=cv2.INTER_AREA)

    video_name = os.path.basename(filepath)
    video_basename_no_suffix_name = os.path.splitext(video_name)[0]

    # os.sep作用：用于系统路径中的分隔符
    #           Windows系统上，文件的路径分隔符是 '\'
    #           Linux系统上，文件的路径分隔符是 '/'
    #           苹果Mac OS系统中是 ':'
    # cv2.imwrite(output_path + os.sep + video_basename_no_suffix_name + '.png', image)
    cv2.imencode('.png', image)[1].tofile(output_path + os.sep + video_basename_no_suffix_name + '.png')

    cap.release()


def create_Excel(filepath):
    """
    用xlsxwriter创建特定格式的Excel文件
    :param filepath: 输出文件路径
    :return:
    """
    # 输出文件路径
    workbook = xlsxwriter.Workbook(filepath)
    # 添加工作表
    worksheet = workbook.add_worksheet()

    # 设置工作表列宽，其中A列为名称，B列为修改日期，C列为大小，D列为时长，E列为图片，图片大小为116像素(20.416cm)*65像素（列宽*行高）
    # worksheet.set_column('A', 40, 'B', 16, 'C', 12, 'D', 8, 'E', 21)
    worksheet.set_column('A:A', 80)
    worksheet.set_column('B:B', 16)
    worksheet.set_column('C:D', 12)
    worksheet.set_column('E:E', 21)

    # 设置默认工作表行高为65像素
    worksheet.set_default_row(65, hide_unused_rows=False)
    worksheet.set_row(row=0, height=15)

    worksheet.write(0, 0, '名称')
    worksheet.write(0, 1, '修改日期')
    worksheet.write(0, 2, '大小')
    worksheet.write(0, 3, '时长')
    worksheet.write(0, 4, '图片')

    workbook.close()


def append_Excel(append_filepath, video_files_path, frame_index):
    """
    在已创建/已完成部分的Excel文档末尾继续添加视频文件信息
    :param append_filepath: 已创建/已完成部分的Excel文档位置
    :param video_files_path: 视频文件位置
    :param frame_index: 生成的缩略图对应的视频帧数
    :return:
    """
    # 读取excel表
    wbook = openpyxl.load_workbook(append_filepath)
    sheet = wbook.active
    # ws = wbook['Sheet1']

    format_ = openpyxl.styles.Alignment(
        # 水平对齐
        horizontal='center',
        # 垂直对齐
        vertical='center',
        # 是否换行，其中参照源码可知：
        #                       wrap_text = Alias('wrapText')
        #                       text_rotation = Alias('textRotation')
        #                       shrink_to_fit = Alias('shrinkToFit')
        wrap_text=True
    )

    currentRow = sheet.max_row

    filenames = os.listdir(video_files_path)

    for i, filename in enumerate(filenames):
        # os.path.splitext可专门分离文件名和后缀名，是对'a.b.c'从后往前遇到第一个分隔符便将分离为'a.b'和'.c'两个部分并返回列表
        portion = os.path.splitext(filename)
        if portion[1] == ".mp4":
            # 发现要写进去的数据则更新至下一行
            currentRow += 1

            modifyTime = get_FileModifyTime(video_files_path + filename)
            size = get_FileSize(video_files_path + filename)
            duration = get_FileDuration(video_files_path + filename)

            minutes, seconds = divmod(duration, 60)
            hours, minutes = divmod(minutes, 60)

            # 设定截取第几帧
            get_SpecifiedFrame(video_files_path + filename, frame_index, r'D:\Downloads\temp')
            image_path = r'D:\Downloads\temp' + os.sep + portion[0] + '.png'
            img = openpyxl.drawing.image.Image(image_path)
            # position = 'E' + currentRow
            # img.anchor(position)
            # 上面方法不行，经查找下面的方法在当前版本可能是唯一可行的
            sheet.add_image(img, anchor=image_column + str(currentRow))

            sheet.cell(row=currentRow, column=1).value = filename
            sheet.cell(row=currentRow, column=1).alignment = format_

            sheet.cell(row=currentRow, column=2).value = modifyTime
            sheet.cell(row=currentRow, column=2).alignment = format_

            sheet.cell(row=currentRow, column=3).value = '%.2f MB' % size
            sheet.cell(row=currentRow, column=3).alignment = format_

            sheet.cell(row=currentRow, column=4).value = '{:d}:{:02d}:{:02d}'.format(hours, minutes, seconds)
            sheet.cell(row=currentRow, column=4).alignment = format_

            print(filename)

    wbook.save(append_filepath)
    shutil.rmtree(r'D:\Downloads\temp')


if __name__ == '__main__':
    # 读取文件路径，一般来说路径过长，因此分成两段
    path1 = "D:/Downloads/压缩包/"
    path2 = "vvv/v_test/"
    video_files_path = path1 + path2

    excel_path = "D:/Downloads/压缩包/list1.xlsx"

    # 如果不存在Excel文件则创建
    if not os.path.isfile(excel_path):
        create_Excel(excel_path)

    append_Excel(excel_path, video_files_path, frame_index=65)
    print("done.")
